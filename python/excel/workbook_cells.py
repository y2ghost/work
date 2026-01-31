from openpyxl import load_workbook


def get_cell_info(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    print(sheet)
    print(f'工作表的标题是: {sheet.title}')
    print(f'值 {sheet["A2"].value=}')
    print(f'值 {sheet["A3"].value=}')
    cell = sheet['B3']
    print(f'值 {cell.value=}')


def get_info_by_coord(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    cell = sheet['A2']
    print(f'行 {cell.row}, 列 {cell.column} = {cell.value}')
    print(f'{cell.value=} 是在 {cell.coordinate=}')


if __name__ == '__main__':
    get_cell_info('books.xlsx')
    get_info_by_coord('books.xlsx')
