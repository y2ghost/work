import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    print(workbook.sheetnames)
    # 添加新的工作表
    workbook.create_sheet()
    print(workbook.sheetnames)
    # 插入新的工作表
    workbook.create_sheet(index=1,
                          title='第二个工作表')
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == '__main__':
    create_worksheets('sheets.xlsx')
