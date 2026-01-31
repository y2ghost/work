"""
这个模块提供读写EXCEL中二维列表数据的函数
"""
import re
import itertools
import datetime as dt

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pyxlsb
except ImportError:
    pyxlsb = None
try:
    import xlrd
    from xlrd.biffh import error_text_from_code
except ImportError:
    xlrd = None
try:
    import xlwt
except ImportError:
    xlwt = None
try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

def read(sheet, first_cell="A1", last_cell=None):
    if xlrd and isinstance(sheet, xlrd.sheet.Sheet):
        if last_cell is None:
            last_cell = (sheet.nrows, sheet.ncols)
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        values = []
        for r in range(first_cell[0] - 1, last_cell[0]):
            row = []
            for c in range(first_cell[1] - 1, last_cell[1]):
                if sheet.cell(r, c).ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        sheet.cell(r, c).value, sheet.book.datemode)
                elif sheet.cell(r, c).ctype in [xlrd.XL_CELL_EMPTY,
                                                xlrd.XL_CELL_BLANK]:
                    value = None
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_ERROR:
                    value = error_text_from_code[sheet.cell(r, c).value]
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(sheet.cell(r, c).value)
                else:
                    value = sheet.cell(r, c).value
                row.append(value)
            values.append(row)
        return values
    elif openpyxl and isinstance(sheet, (openpyxl.worksheet.worksheet.Worksheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet)):
        if last_cell is None:
            last_cell = (sheet.max_row, sheet.max_column)
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.cell.coordinate_to_tuple(first_cell)
        if not isinstance(last_cell, tuple):
            last_cell = openpyxl.utils.cell.coordinate_to_tuple(last_cell)
        data = []
        for row in sheet.iter_rows(min_row=first_cell[0], min_col=first_cell[1],
            max_row=last_cell[0], max_col=last_cell[1], values_only=True):
            data.append(list(row))
        return data
    elif pyxlsb and isinstance(sheet, pyxlsb.worksheet.Worksheet):
        errors = {"0x0": "#NULL!", "0x7": "#DIV/0!", "0xf": "#VALUE!",
            "0x17": "#REF!", "0x1d": "#NAME?", "0x24": "#NUM!", "0x2a": "#N/A"}
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if last_cell and not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        data = []
        for row in itertools.islice(sheet.rows(),
            first_cell[0] - 1, last_cell[0] if last_cell else None):
            data.append([errors.get(cell.v, cell.v) for cell in row]
                        [first_cell[1] - 1 : last_cell[1] if last_cell else None])
        return data
    else:
        raise TypeError(f"无法处理{type(sheet)}类型的工作簿")

def write(sheet, values, first_cell="A1", date_format=None):
    if openpyxl and isinstance( sheet, openpyxl.worksheet.worksheet.Worksheet):
        if date_format is None:
            date_format = "mm/dd/yy"
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.coordinate_to_tuple(first_cell)
        for i, row in enumerate(values):
            for j, value in enumerate(row):
                cell = sheet.cell(row=first_cell[0] + i, column=first_cell[1] + j)
                cell.value = value
                if date_format and isinstance(value, (dt.datetime, dt.date)):
                    cell.number_format = date_format
    elif xlsxwriter and isinstance(sheet, xlsxwriter.worksheet.Worksheet):
        if date_format is not None:
            raise ValueError("暂不支持date_format参数，请在sheet对象上设置")
        if isinstance(first_cell, tuple):
            first_cell = first_cell[0] - 1, first_cell[1] - 1
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for r, row_data in enumerate(values):
            sheet.write_row(first_cell[0] + r, first_cell[1], row_data)
    elif xlwt and isinstance(sheet, xlwt.Worksheet):
        if date_format is None:
            date_format = "mm/dd/yy"
        date_format = xlwt.easyxf(num_format_str=date_format)
        if isinstance(first_cell, tuple):
            first_cell = (first_cell[0] - 1, first_cell[1] - 1)
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for i, row in enumerate(values):
            for j, cell in enumerate(row):
                if isinstance(cell, (dt.datetime, dt.date)):
                    sheet.write(i + first_cell[0], j + first_cell[1], cell, date_format)
                else:
                    sheet.write(i + first_cell[0], j + first_cell[1], cell)
    else:
        raise TypeError(f"不能处理{type(sheet)}类型的工作簿")

def xl_cell_to_rowcol(cell_str):
    if not cell_str:
        return 0, 0

    match = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)").match(cell_str)
    col_str = match.group(2)
    row_str = match.group(4)
    expn = 0
    col = 0

    for char in reversed(col_str):
        col += (ord(char) - ord("A") + 1) * (26 ** expn)
        expn += 1

    row = int(row_str) - 1
    col -= 1
    return row, col
