import openpyxl
import excel
import datetime as dt

book = openpyxl.load_workbook("xl/stores.xlsx", data_only=True)
# 演示sheet获取的两种方法
sheet = book["2019"]
sheet = book.worksheets[0]
print(book.sheetnames)

for i in book.worksheets:
    print(f"titile: {i.title}")

print(f"sheet.max_row: {sheet.max_row}, sheet.max_column: {sheet.max_column}")
print(f'sheet["B6"]= {sheet["B6"].value}')
print(f'sheet.cell(row=6, column=2) = {sheet.cell(row=6, column=2).value}')

data = excel.read(sheet, (2, 2), (8, 6))
print(f"data[:2] = {data[:2]}")
