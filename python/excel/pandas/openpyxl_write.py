import excel
import openpyxl
import datetime as dt
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill

book = openpyxl.Workbook()
sheet = book.active
sheet.title = "test"

sheet["A1"].value = "你好1"
sheet.cell(row=2, column=1, value = "你好2")

font_format = Font(color="FF0000", bold=True)
thin = Side(border_style="thin", color="FF0000")

# 应用字体和y样式
sheet["A3"].value = "你好3"
sheet["A3"].font = font_format
sheet["A3"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet["A3"].alignment = Alignment(horizontal="center")
sheet["A3"].fill = PatternFill(fgColor="FFFF00", fill_type="solid")

# 数字格式
sheet["A4"].value = 3.3333
sheet["A4"].number_format = "0.00"

# 日期格式
sheet["A5"].value = dt.date(2016, 10, 13)
sheet["A5"].number_format = "mm/dd/yy"

# 公式示例
sheet["A6"].value = "=SUM(A4, 2)"

# 插入图片
sheet.add_image(Image("images/python.png"), "C1")

# 二维列表操作示例
data = [[None, "North", "South"], ["Last Year", 2, 5], ["This Year", 3, 6]]
excel.write(sheet, data, "A10")

# 生成图表示例
chart = BarChart()
chart.type = "col"
chart.title = "Sales Per Region"
chart.x_axis.title = "Regions"
chart.y_axis.title = "Sales"
chart_data = Reference(sheet, min_row=11, min_col=1,
    max_row=12, max_col=3)
chart_categories = Reference(sheet, min_row=10, min_col=2,
    max_row=10, max_col=3)
chart.add_data(chart_data, titles_from_data=True, from_rows=True)
chart.set_categories(chart_categories)
sheet.add_chart(chart, "A15")

# 保存文件
book.save("openpyxl.xlsx")