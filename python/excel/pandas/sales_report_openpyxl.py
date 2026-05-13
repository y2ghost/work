import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

this_dir = Path(__file__).resolve().parent
parts = []

for path in (this_dir / "sales_data").rglob("*.xls*"):
    print(f'Reading {path.name}')
    part = pd.read_excel(path)
    parts.append(part)

df = pd.concat(parts)
pivot = pd.pivot_table(df, index="transaction_date", columns="store",
    values="amount", aggfunc="sum")
summary = pivot.resample("M").sum()
summary.index.name = "Month"
summary = summary.loc[:, summary.sum().sort_values().index]
summary.loc[:, "Total"] = summary.sum(axis=1)
summary.sum(axis=0).rename("Total")

# 输出数据到excel文件
startrow, startcol = 3, 2
nrows, ncols = summary.shape

with pd.ExcelWriter(this_dir / "sales_report_openpyxl.xlsx", engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Sheet1", startrow=startrow - 1, startcol=startcol - 1)
    book = writer.book
    sheet = writer.sheets["Sheet1"]
    sheet.cell(row=1, column=startcol, value="Sales Report")
    sheet.cell(row=1, column=startcol).font = Font(size=24, bold=True)
    sheet.sheet_view.showGridLines = False

    for row in range(startrow + 1, startrow + nrows + 1):
        for col in range(startcol + 1, startcol + ncols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")

    for cell in sheet["B"]:
        cell.number_format = "mmm yy"

    for col in range(startcol, startcol + ncols + 1):
        cell = sheet.cell(row=startrow, column=col)
        sheet.column_dimensions[cell.column_letter].width = 14

    first_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
    last_cell = sheet.cell(row=startrow + nrows, column=startcol + ncols)
    range_address = f"{first_cell.coordinate}:{last_cell.coordinate}"
    sheet.conditional_formatting.add(range_address, CellIsRule(operator="lessThan", formula=["20000"], stopIfTrue=True, font=Font(color="E93423")))

    chart = BarChart()
    chart.type = "col"
    chart.title = "Sales per Month and Store"
    chart.height = 11.5
    chart.width = 20.5

    data = Reference(sheet, min_col=startcol + 1, min_row=startrow, max_row=startrow + nrows - 1, max_col=startcol + ncols - 1)
    categories = Reference(sheet, min_col=startcol, min_row=startrow + 1, max_row=startrow + nrows - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    cell = sheet.cell(row=startrow + nrows + 2, column=startcol)
    sheet.add_chart(chart=chart, anchor=cell.coordinate)

    chart.y_axis.title = "Sales"
    chart.x_axis.title = summary.index.name
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
