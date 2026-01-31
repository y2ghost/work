from openpyxl import load_workbook


def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f'工作表名称: {workbook.sheetnames}')
    sheet = workbook.active
    print(sheet)
    print(f'工作表的标题是: {sheet.title}')


if __name__ == '__main__':
    open_workbook('books.xlsx')
